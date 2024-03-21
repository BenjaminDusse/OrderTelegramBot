import os
import sys
import time
import sqlite3
import traceback
from datetime import datetime
from openpyxl import Workbook
from app.config import db_name_1

app_exit = False


def create_db_table(dblink):
    with sqlite3.connect(dblink) as base:
        print("Checking db")
        cursor = base.cursor()
        cursor.execute('''CREATE TABLE IF NOT EXISTS users ( 
                user_id INTEGER PRIMARY KEY,
                username TEXT NOT NULL)''')

        cursor.execute('''
            CREATE TABLE IF NOT EXISTS Inventory (
                inv_id INT PRIMARY KEY,
                start_date DATETIME,
                end_date DATETIME,
                status TEXT DEFAULT 'ongoing' CHECK(status IN ('ongoing', 'completed')))
        ''')

        cursor.execute('''
            CREATE TABLE IF NOT EXISTS items (
                item_id INTEGER PRIMARY KEY AUTOINCREMENT,
                code VARCHAR(5),
                title TEXT,
                price INTEGER,
                amount REAL)
        ''')

        cursor.execute('''
            CREATE TABLE IF NOT EXISTS barcode(
                barcode INTEGER PRIMARY KEY, 
                code INTEGER)
        ''')

        # Adding a user_id column to the UserDocument table definition
        cursor.execute('''
        CREATE TABLE IF NOT EXISTS UserDocument (
            user_document_id INTEGER PRIMARY KEY,
            user_id INTEGER NOT NULL,
            inv_id INTEGER NOT NULL,
            item_id INTEGER NOT NULL,
            barcode INTEGER NOT NULL,
            amount REAL,
            FOREIGN KEY (inv_id) REFERENCES inventory(inv_id),
            FOREIGN KEY (item_id) REFERENCES items(item_id),
            FOREIGN KEY (user_id) REFERENCES users(user_id),
            FOREIGN KEY (barcode) REFERENCES barcode(barcode))
        ''')


def db_connect(dblink):
    if os.path.exists(dblink) is False:
        create_db_table(dblink)
    with sqlite3.connect(dblink) as base:
        cursor = base.cursor()
        return base, cursor


def create_inventory(base, cursor, generated_num, current_datetime=datetime.now()):
    cursor.execute(
        "INSERT INTO Inventory(inv_id, start_date) VALUES(?, ?)", (
            generated_num, current_datetime)
    )
    base.commit()


def get_ongoing_inventory(cursor):
    cursor.execute(
        '''
            select
                inv_id
            from
                Inventory
            where
                status = 'ongoing'
        '''
    )
    inv_id = cursor.fetchone()
    if inv_id:
        return inv_id[0]
    return None


def update_ongoing_inventory(
        base,
        cursor,
        inv_id,
        end_date=datetime.now()
):
    cursor.execute(
        '''
            UPDATE Inventory
            SET status = 'completed', end_date = ?
            where
                inv_id = ? 
        ''', (end_date, inv_id)
    )
    base.commit()

    # print(user_id, inv_id, barcode, new_amount)
    # cursor.execute(
    #     '''
    #         UPDATE UserDocument
    #         SET amount = amount + ?
    #         WHERE
    #             user_id = ? AND
    #             inv_id = (SELECT inv_id FROM inventory WHERE inv_id = ?) AND
    #             barcode = (SELECT barcode FROM barcode WHERE barcode = ?)
    #     ''', (new_amount, user_id, inv_id, barcode)
    # )
    #
    # # Commit the transaction to save the changes
    # base.commit()


def fetch_product_by_barcode(base, cursor, barcode):
    cursor.execute(
        '''
        SELECT 
            i.title,
            i.amount
        FROM 
            items i 
        JOIN barcode b ON i.code = b.code
        WHERE
            b.barcode = ?''', (barcode,))
    product = cursor.fetchone()
    base.commit()
    return product


def fetch_document(base, cursor, user_id, inv_id, barcode):
    cursor.execute(
        '''
            SELECT
                i.title,
                doc.amount
            FROM
                UserDocument doc
            JOIN
                items i ON doc.item_id = i.item_id
            JOIN
                inventory inv ON doc.inv_id = inv.inv_id
            JOIN
                barcode b ON doc.barcode = b.barcode
            WHERE
                doc.user_id = ? AND inv.inv_id = ? and b.barcode = ?
        ''', (user_id, inv_id, barcode)
    )
    document = cursor.fetchone()
    base.commit()
    return document


def update_product_amount(base, cursor, inv_id, item_id, user_id, barcode, amount):
    cursor.execute(
        "INSERT INTO UserDocument(inv_id, item_id, user_id, barcode, amount) VALUES(?,?,?,?,?)", (
            inv_id, item_id, user_id, barcode, amount)
    )

    base.commit()


def update_document_amount(base, cursor, user_id, inv_id, barcode, new_amount):
    # Prepare the UPDATE statement to change the doc.amount
    print(user_id, inv_id, barcode, new_amount)
    cursor.execute(
        '''
            UPDATE UserDocument
            SET amount = amount + ?
            WHERE
                user_id = ? AND
                inv_id = (SELECT inv_id FROM inventory WHERE inv_id = ?) AND
                barcode = (SELECT barcode FROM barcode WHERE barcode = ?)
        ''', (new_amount, user_id, inv_id, barcode)
    )

    # Commit the transaction to save the changes
    base.commit()


def fetch_product(base, cursor, product_code):
    cursor.execute(
        '''
    SELECT 
        i.item_id,
        i.code,
        i.title,
        i.price,
        i.amount as "plan_amount",
        b.barcode
    FROM 
        items i
    JOIN barcode b ON i.code = b.code
    WHERE 
        b.barcode = ?''', (product_code,))
    product = cursor.fetchone()
    base.commit()
    return product


# async def fetch_document_for_inventory_report(base, cursor, message, inv_id):
#     cursor.execute('SELECT user_id, username FROM users')
#     user_ids = cursor.fetchall()
#     print(user_ids)
#
#     for user_id in user_ids:
#         print("In loop")
#         print(user_id, user_id[0], user_id[1])
#         user_id, username = str(user_id[0]), user_id[1]
#         print("After getting")
#         print(user_id, username)
#         wb = Workbook()
#         ws = wb.active
#         column_names = ["Inventarizatsiya raqami", "Tovar kodi", "Tovar nomi", "Narx",
#                         "Miqdori(plan)", "Produkt shtrix kodi", "Miqdori(fakt)"]
#         ws.append(column_names)
#
#         # Your existing query modified to use string formatting for user_id and inv_id
#         cursor.execute(
#             '''
#             SELECT
#                 doc.inv_id,
#                 doc.item_id,
#                 i.title,
#                 i.price,
#                 i.amount,
#                 b.barcode,
#                 doc.amount
#             FROM
#                 UserDocument doc
#             JOIN
#                 items i ON doc.item_id = i.item_id
#             JOIN
#                 inventory inv ON doc.inv_id = inv.inv_id
#             JOIN
#                 barcode b ON doc.barcode = b.barcode
#             WHERE
#                 doc.user_id = ? AND inv.inv_id = ?
#             ''', (user_id, inv_id)
#         )
#
#         documents = cursor.fetchall()
#         if documents:
#             for document in documents:
#                 ws.append(document)
#
#             # Save the workbook with filename as user_id
#             filename = f"{username}_{user_id}.xlsx"
#             wb.save(filename)
#             await message.answer_document(open(filename, 'rb'))
#             os.remove(filename)
#
#         base.commit()
#     return user_ids

async def fetch_document_for_inventory_report(base, cursor, message, inv_id):
    wb = Workbook()
    ws = wb.active
    column_names = ["Inventarizatsiya raqami", "Tovar kodi", "Tovar nomi", "Narx",
                    "Miqdori(plan)", "Produkt shtrix kodi", "Miqdori(fakt)"]
    ws.append(column_names)

    cursor.execute(
        '''
        SELECT
            doc.inv_id,
            doc.item_id,
            i.title,
            i.price,    
            i.amount,
            b.barcode,
            doc.amount
        FROM
            UserDocument doc
        JOIN
            items i ON doc.item_id = i.item_id
        JOIN
            inventory inv ON doc.inv_id = inv.inv_id
        JOIN
            barcode b ON doc.barcode = b.barcode
        WHERE
            inv.inv_id = ?
        ''', (inv_id,)
    )

    documents = cursor.fetchall()
    if documents:
        for document in documents:
            ws.append(document)

        # Save the workbook with filename as user_id
        current_time = datetime.now().strftime('%d.%m.%Y')
        filename = f"{current_time}_inventarizatsiya.xlsx"
        wb.save(filename)
        await message.answer_document(open(filename, 'rb'))
        os.remove(filename)

    base.commit()


def delete_database(database_name):
    try:
        os.remove(database_name)
        print(f"Database '{database_name}' deleted successfully.")
    except FileNotFoundError:
        print(f"Database '{database_name}' not found.")
    except Exception as e:
        print(f"Error deleting database '{database_name}': {e}")


async def fetch_inventory_data(cursor, inv_id):
    cursor.execute(
        '''
            SELECT
                doc.inv_id as "inventory_id",
                doc.item_id as "kod produkta",
                i.title as "produkt nomi",
                b.barcode as "produkt_shtrix_kodi", 
                i.price as "produkt narxi",
                i.amount as "amount_plan",
                doc.amount as "amount_fact",
                i.amount * i.price as  "summ_plan",
                doc.amount * i.price as "summ_fakt",
                doc.amount - i.amount  "raznitsa amount",
                (doc.amount * i.price) - (i.amount * i.price) as "raznitsa fakt"
            FROM
                UserDocument doc
            JOIN
                items i ON doc.item_id = i.item_id
            JOIN
                inventory inv ON doc.inv_id = inv.inv_id
            JOIN
                barcode b ON doc.barcode = b.barcode
            WHERE
                inv.inv_id = ?
        ''', (inv_id,)
    )
    results = cursor.fetchall()

    return results


async def generate_excel_file(data, file_name):
    wb = Workbook()
    ws = wb.active
    column_names = [
        "Inventory ID",
        "Produkt kodi",
        "Produkt nomi",
        "Produkt shtrix kodi",
        "Produkt narxi",
        "Boshlang'ich miqdori",
        "Kiritilgan miqdori",
        "Boshlang'ich summa",
        "Kiritilgan summa",
        "Raznitsa amount",
        "Raznitsa amount",
        "User_id"
    ]

    # Add the column names to the first row
    ws.append(column_names)

    for row in data:
        ws.append(row)
    wb.save(filename=file_name)
    return file_name


def create_document_for_product(base, cursor, inv_id, item_id, user_id, barcode, product_code, amount):
    _, _, _, _, product_amount, _ = fetch_product(base, cursor, product_code)

    cursor.execute(
        "INSERT INTO UserDocument(inv_id, item_id, user_id, barcode, amount) VALUES(?,?,?,?,?)", (
            inv_id, item_id, user_id, barcode, amount)
    )

    base.commit()


def get_user(base, cursor, message):
    cursor.execute(
        "SELECT user_id, username FROM users where user_id = ?", (
            message.from_user.id,)
    )
    user = cursor.fetchone()
    base.commit()
    return user


def create_user(base, cursor, message):
    cursor.execute(
        "INSERT INTO users(user_id, username) VALUES(?,?)", (message.from_user.id, message.from_user.username))
    base.commit()


def insert_item(base, cursor, code: int, title: str, price: float, amount: float):
    cursor.execute(
        '''INSERT OR REPLACE INTO items(code, title, price, amount) VALUES(?,?,?,?)''', (code, title, price, amount))
    base.commit()


def insert_barcode(base, cursor, code: int, barcode: str):
    try:
        cursor.execute(
            '''INSERT OR REPLACE INTO barcode(barcode, code) VALUES(?,?)''', (barcode, code,))
        base.commit()
    except Exception as ex:
        print(ex)
        print(code, barcode)


def atol_pars(dblink: str, atol_txt: str):
    try:
        with open(atol_txt, mode='r', encoding='cp1251') as f:
            file = f.read().encode().decode("utf-8")
        data = file.split('\n')
        item_start_index, item_stop_index = None, None
        barcode_start_index, barcode_stop_index = None, None

        try:
            item_start_index = data.index("$$$REPLACEQUANTITY") + 1
            item_stop_index = data[item_start_index:].index("") + item_start_index
        except Exception as ex:
            print(ex)
        try:
            barcode_start_index = data.index("$$$ADDBARCODES") + 1
            barcode_stop_index = data[barcode_start_index:].index(
                "") + barcode_start_index
        except Exception as ex:
            print(ex)

        base, cursor = db_connect(dblink)
        if item_start_index is not None:
            for item in data[item_start_index:item_stop_index]:
                item_split = item.split(';')
                code, name, price, amount = item_split[0], item_split[2], item_split[4], item_split[5]
                if not amount:
                    amount = 0
                amount_str = str(amount)
                amount = float(amount_str.replace(',', '.'))

                if price.isdigit():
                    if float(price) > 0:
                        insert_item(base, cursor, int(code),
                                    name, float(price), amount)

        if barcode_start_index is not None:
            for barcode in data[barcode_start_index:barcode_stop_index]:
                barcode_split = barcode.split(';')
                code, bc = barcode_split[1], barcode_split[0]
                insert_barcode(base, cursor, int(code), bc)
    except Exception as ex:
        print(ex)
        print(traceback.format_exc())


def update_start(db_name: str, input_file: str):
    print("In function")

    try:
        if app_exit is True:
            sys.exit()

        if os.path.exists(input_file):
            print('---Есть загрузка---')
            print(f"Started at {datetime.now().strftime('%H:%M')}")
            atol_pars(db_name, input_file)
            print(f"End in {datetime.now().strftime('%H:%M')}")
            print(f'--- Загрузка выполнено! ---')
            os.remove(input_file)
        time.sleep(2)
    except Exception as ex:
        print(ex)
        print(traceback.format_exc())
        time.sleep(10)


def update_stop():
    global app_exit
    app_exit = True


inventory_session = {
    "active": False,
    "db_filled": False,
    "inv_id": 0,
}

base_1, cursor_1 = db_connect(db_name_1)
print(db_name_1)
